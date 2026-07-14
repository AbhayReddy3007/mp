#!/usr/bin/env python3
"""
Clinical Evidence Analyzer v3
==============================
6-step pharmaceutical analyst workflow with PDGEC scoring
and BigQuery-ready flat table output.

Usage:
    python clinical_evidence.py --molecule semaglutide
    python clinical_evidence.py --molecule dupilumab --max-trials 100
    python clinical_evidence.py --molecule semaglutide liraglutide tirzepatide

Requires:  .env with BQ_SERVICE_ACCOUNT=/path/to/key.json

Concurrency notes:
    - Multiple molecules run in parallel (ThreadPoolExecutor, main()).
    - Each molecule's log output is buffered in a contextvar-backed
      StringIO and flushed atomically (single print call) when that
      molecule finishes, so parallel runs don't interleave their
      console output line-by-line.
    - All calls to the Gemini API (across all molecules and all their
      inner worker threads) are gated by a single global semaphore
      (GEMINI_MAX_CONCURRENT, default 5) to avoid bursts of concurrent
      requests tripping rate limits when many molecules/sources fire
      at once.
"""

import argparse
import contextvars
import csv
import io
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from collections import OrderedDict
from statistics import mode as _stat_mode

import requests
from dotenv import load_dotenv

try:
    import google.generativeai as genai
    _GENAI_AVAILABLE = True
except ImportError:
    _GENAI_AVAILABLE = False

# -------------------------------------------------------------------
# Config
# -------------------------------------------------------------------
load_dotenv()
BQ_SERVICE_ACCOUNT = os.getenv("BQ_SERVICE_ACCOUNT")
GEMINI_API_KEY     = os.getenv("GEMINI_API_KEY", "").strip()

CTGOV_API      = "https://clinicaltrials.gov/api/v2/studies"
FDA_API        = "https://api.fda.gov/drug/drugsfda.json"
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

# -------------------------------------------------------------------
# Global Gemini concurrency cap
# -------------------------------------------------------------------
# All Gemini calls -- across every molecule being processed in parallel,
# and across every inner worker thread (RWE sources, E/C scoring) --
# acquire this semaphore first. This bounds total simultaneous requests
# to the Gemini API regardless of how many molecules/threads are running,
# which is what actually protects against rate-limit (429) errors.
GEMINI_MAX_CONCURRENT = int(os.getenv("GEMINI_MAX_CONCURRENT", "5"))
_gemini_semaphore = threading.Semaphore(GEMINI_MAX_CONCURRENT)

# Initialise google-generativeai client (used for RWE summariser + scoring)
_gemini_model = None
if not _GENAI_AVAILABLE:
    print("[WARN] google-generativeai not installed — LLM scoring disabled. "
          "Run: pip install google-generativeai")
elif not GEMINI_API_KEY:
    print("[WARN] GEMINI_API_KEY not set in .env — LLM scoring disabled. "
          "Add GEMINI_API_KEY=<your key> to your .env file.")
else:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        _gemini_model = genai.GenerativeModel("gemini-2.5-flash")
        print(f"[OK]  Gemini 2.5 Flash initialised (key: ...{GEMINI_API_KEY[-6:]}) "
              f"| max concurrent calls: {GEMINI_MAX_CONCURRENT}")
    except Exception as e:
        print(f"[ERR] Gemini model init failed: {e} — falling back to rule-based scoring.")
        _gemini_model = None


# -------------------------------------------------------------------
# Logger (ASCII only for Windows compat)
# -------------------------------------------------------------------
# Per-molecule log buffering
# -------------------------------------------------------------------
# When multiple molecules run concurrently (see main()), each molecule's
# worker thread sets this contextvar to its own io.StringIO buffer before
# doing any work. Because contextvars.copy_context() is used whenever a
# NEW worker thread is spawned from within that molecule's processing
# (see submit_with_ctx below), the buffer reference correctly propagates
# down into every nested thread pool (trial-source fetches, RWE-source
# fetches, etc.) that belongs to that molecule.
#
# All Log.* calls check this contextvar: if a buffer is present, the log
# line is appended to it instead of being printed immediately. The caller
# (main()._process_molecule) then prints the entire buffer in a single
# print() call once the molecule is fully done, so one molecule's log
# output can never be interleaved line-by-line with another's.
#
# If no buffer is set (e.g. single-molecule runs, or code running outside
# a molecule context), Log falls back to printing directly, so behavior
# is unchanged from before for non-concurrent use.
_log_buffer_var = contextvars.ContextVar("log_buffer", default=None)


def submit_with_ctx(executor, fn, *args, **kwargs):
    """
    Submit fn to executor, explicitly copying the current contextvars
    context (including _log_buffer_var) into the worker thread.

    Plain ThreadPoolExecutor.submit() does NOT propagate contextvars to
    the new OS thread -- each new thread starts with a fresh/default
    context. Without this helper, nested thread pools (e.g. the trial
    source fetches or RWE source fetches spawned inside run_molecule)
    would lose track of which molecule's log buffer they belong to and
    fall back to printing directly, reintroducing interleaved output.
    """
    ctx = contextvars.copy_context()
    return executor.submit(ctx.run, fn, *args, **kwargs)


class Log:
    @staticmethod
    def _emit(line):
        buf = _log_buffer_var.get()
        if buf is not None:
            buf.write(line + "\n")
        else:
            print(line)

    @staticmethod
    def header(msg):  Log._emit(f"\n{'='*72}\n  {msg}\n{'='*72}")
    @staticmethod
    def step(n, msg): Log._emit(f"\n{'-'*72}\n  STEP {n}: {msg}\n{'-'*72}")
    @staticmethod
    def info(msg):    Log._emit(f"    {msg}")
    @staticmethod
    def ok(msg):       Log._emit(f"    [OK]   {msg}")
    @staticmethod
    def warn(msg):     Log._emit(f"    [WARN] {msg}")
    @staticmethod
    def err(msg):
        buf = _log_buffer_var.get()
        line = f"    [ERR]  {msg}"
        if buf is not None:
            buf.write(line + "\n")
        else:
            print(line, file=sys.stderr)
    @staticmethod
    def sub(msg):      Log._emit(f"\n  --- {msg} ---")
    @staticmethod
    def raw(text):
        """Emit a large pre-formatted block (e.g. the final report) as-is."""
        buf = _log_buffer_var.get()
        if buf is not None:
            buf.write(text + "\n")
        else:
            print(text)

log = Log()


def http_get(url, params=None, retries=3, timeout=30):
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, timeout=timeout)
            if r.status_code == 200:
                return r
            if r.status_code == 429:
                time.sleep(2 ** (attempt + 1)); continue
            log.warn(f"HTTP {r.status_code} from {url}")
            return r
        except requests.RequestException as e:
            log.warn(f"Request failed (attempt {attempt+1}): {e}")
            time.sleep(2)
    return None


def safe_str(val):
    """Convert any value to an ASCII-safe string for CSV/report output."""
    if val is None:
        return ""
    s = str(val)
    # Replace non-ASCII with closest ASCII or strip
    return s.encode("ascii", errors="replace").decode("ascii")


def _wrap_text(text, width=88, indent="    "):
    """Wrap long text to specified width with indent prefix."""
    import textwrap
    lines = textwrap.wrap(text, width=width - len(indent))
    return "\n".join(f"{indent}{line}" for line in lines) if lines else f"{indent}{text}"


# ===================================================================
# STEP 0 -- Molecule Normalization (fully dynamic from web sources)
# ===================================================================

def _fetch_brand_names_from_fda(name: str) -> tuple:
    """Fetch brand names, manufacturer, and dev codes from openFDA."""
    brand_names = set()
    manufacturers = set()
    app_numbers = set()

    for field in ["generic_name", "brand_name", "substance_name"]:
        r = http_get(FDA_API, params={
            "search": f'openfda.{field}:"{name}"', "limit": 20,
        })
        if r and r.status_code == 200:
            try:
                for res in r.json().get("results", []):
                    openfda = res.get("openfda", {})
                    for bn in openfda.get("brand_name", []):
                        brand_names.add(bn)
                    for mfg in openfda.get("manufacturer_name", []):
                        manufacturers.add(mfg)
                    app_no = res.get("application_number", "")
                    if app_no:
                        app_numbers.add(app_no)
            except Exception:
                pass

    return list(brand_names), list(manufacturers), list(app_numbers)


def _fetch_info_from_ctgov(name: str) -> dict:
    """Fetch sponsor, dev codes, indications from ClinicalTrials.gov."""
    info = {"sponsor": "Unknown", "dev_codes": [], "indications": [],
            "other_ids": []}
    try:
        r = http_get(CTGOV_API, params={
            "query.intr": name, "pageSize": 20, "format": "json",
        })
        if not r or r.status_code != 200:
            return info

        studies = r.json().get("studies", [])
        if not studies:
            return info

        # Sponsor from first study
        sp = (studies[0].get("protocolSection", {})
              .get("sponsorCollaboratorsModule", {})
              .get("leadSponsor", {}).get("name", "Unknown"))
        info["sponsor"] = sp

        # Collect indications + other IDs (potential dev codes)
        indications = set()
        other_ids = set()
        for s in studies[:20]:
            proto = s.get("protocolSection", {})
            # Conditions
            conds = proto.get("conditionsModule", {}).get("conditions", [])
            indications.update(c.lower() for c in conds)
            # Secondary IDs (dev codes like NN9535, LY3298176, etc.)
            id_mod = proto.get("identificationModule", {})
            for sec_id in id_mod.get("secondaryIdInfos", []):
                sid = sec_id.get("id", "")
                if sid and not sid.startswith("NCT"):
                    other_ids.add(sid)
            org_id = id_mod.get("orgStudyIdInfo", {}).get("id", "")
            if org_id and not org_id.startswith("NCT"):
                other_ids.add(org_id)

        info["indications"] = list(indications)[:15]
        info["dev_codes"] = list(other_ids)[:10]

    except Exception as e:
        log.warn(f"ClinicalTrials.gov discovery error: {e}")

    return info


_COMBO_SEPARATOR_PATTERN = re.compile(r"\s*[+/]\s*")


def _split_combo_components(name: str) -> list:
    """
    Split a combination-therapy name into its individual active-ingredient
    components. Recognizes '+' and '/' as separators, e.g.:
        "Cagrilintide+Semaglutide" -> ["Cagrilintide", "Semaglutide"]
        "cagrilintide/semaglutide" -> ["cagrilintide", "semaglutide"]
    Returns a single-item list (just the original, trimmed) for
    non-combination names.
    """
    parts = [p.strip() for p in _COMBO_SEPARATOR_PATTERN.split(name) if p.strip()]
    return parts if len(parts) > 1 else [name.strip()]


def _combo_name_variants(components: list) -> list:
    """
    Given ["Cagrilintide", "Semaglutide"], generate the alternate ways a
    combination therapy's name commonly appears across FDA/registry
    sources, so lookups aren't limited to a single '+'-joined spelling.
    """
    variants = set()
    if len(components) < 2:
        return []
    variants.add("+".join(components))
    variants.add("/".join(components))
    variants.add(" and ".join(components))
    variants.add(" ".join(components))
    # Reversed order too -- registries aren't consistent about which
    # active ingredient is listed first.
    rev = list(reversed(components))
    variants.add("+".join(rev))
    variants.add("/".join(rev))
    variants.add(" and ".join(rev))
    variants.add(" ".join(rev))
    return sorted(variants)


def normalize_molecule(name: str) -> dict:
    """
    Step 0: Build a normalized molecule record entirely from web sources.

    Handles combination therapies (e.g. "Cagrilintide+Semaglutide") by
    detecting the '+' / '/' separator, looking up each active ingredient
    individually on top of the combined name, and merging the results.
    Single-ingredient molecules ("Semaglutide", "Orforglipron Calcium")
    go through the same path unchanged -- _split_combo_components just
    returns a one-item list for them, so nothing extra happens.
    """
    key = name.strip().lower()
    components = _split_combo_components(name)
    is_combo = len(components) > 1

    record = {
        "generic_name": name.strip().title(),
        "brand_names": [],
        "dev_codes": [],
        "sponsor": "Unknown",
        "indications": [],
        "synonyms": [key],
        "is_combination": is_combo,
        "components": components if is_combo else [],
    }

    if is_combo:
        log.info(f"'{name}' detected as a combination therapy: "
                  f"{' + '.join(components)}")
        # Seed synonyms with alternate separator spellings so downstream
        # BigQuery LIKE-matching and search queries aren't limited to the
        # exact '+'-joined string the user typed.
        record["synonyms"].extend(v.lower() for v in _combo_name_variants(components))

    # 1) Fetch brand names + manufacturer from openFDA
    log.info(f"Looking up '{name}' on openFDA...")
    brands, manufacturers, app_nos = _fetch_brand_names_from_fda(name)

    # 2) Fetch sponsor, dev codes, indications from ClinicalTrials.gov
    log.info(f"Looking up '{name}' on ClinicalTrials.gov...")
    ctgov_info = _fetch_info_from_ctgov(name)

    if is_combo:
        # The exact combo string often isn't registered verbatim anywhere,
        # so also look up each individual active ingredient and merge in
        # anything the combined-name lookup didn't already find. This is
        # what actually makes combo names like "Cagrilintide+Semaglutide"
        # return useful sponsor/brand/indication data instead of coming
        # back empty.
        combo_brands, combo_mfgs = set(brands), set(manufacturers)
        combo_indications = set(ctgov_info["indications"])
        combo_dev_codes = set(ctgov_info["dev_codes"])
        combo_sponsor = ctgov_info["sponsor"]

        for component in components:
            log.info(f"  Looking up component '{component}' on openFDA...")
            c_brands, c_mfgs, _ = _fetch_brand_names_from_fda(component)
            combo_brands.update(c_brands)
            combo_mfgs.update(c_mfgs)
            record["synonyms"].append(component.lower())

            log.info(f"  Looking up component '{component}' on ClinicalTrials.gov...")
            c_ctgov = _fetch_info_from_ctgov(component)
            combo_indications.update(c_ctgov["indications"])
            combo_dev_codes.update(c_ctgov["dev_codes"])
            record["synonyms"].extend(d.lower() for d in c_ctgov["dev_codes"])
            if combo_sponsor == "Unknown" and c_ctgov["sponsor"] != "Unknown":
                combo_sponsor = c_ctgov["sponsor"]

        brands = list(combo_brands)
        manufacturers = list(combo_mfgs)
        ctgov_info["indications"] = list(combo_indications)[:15]
        ctgov_info["dev_codes"] = list(combo_dev_codes)[:10]
        ctgov_info["sponsor"] = combo_sponsor
        if combo_brands:
            log.ok(f"Combined component brand names: "
                   f"{', '.join(sorted(combo_brands)[:5])}")

    if brands:
        record["brand_names"] = sorted(set(b for b in brands
                                           if b.lower() != key))[:10]
        record["synonyms"].extend(b.lower() for b in record["brand_names"])
        log.ok(f"FDA brand names: {', '.join(record['brand_names'][:5])}")
    if manufacturers:
        # Use the most common manufacturer as sponsor
        record["sponsor"] = manufacturers[0]
        log.ok(f"FDA manufacturer: {record['sponsor']}")

    # ClinicalTrials.gov sponsor overrides FDA manufacturer if available
    if ctgov_info["sponsor"] != "Unknown":
        record["sponsor"] = ctgov_info["sponsor"]
        log.ok(f"CTgov sponsor: {record['sponsor']}")

    if ctgov_info["dev_codes"]:
        record["dev_codes"] = ctgov_info["dev_codes"]
        record["synonyms"].extend(d.lower() for d in record["dev_codes"])
        log.ok(f"Dev codes: {', '.join(record['dev_codes'][:5])}")

    if ctgov_info["indications"]:
        record["indications"] = ctgov_info["indications"]
        log.ok(f"Indications: {', '.join(record['indications'][:5])}")

    # Deduplicate synonyms
    record["synonyms"] = list(OrderedDict.fromkeys(record["synonyms"]))

    return record


# ===================================================================
# STEP 1 -- Global Clinical Trial Evidence Scan
# ===================================================================

# -- 1A: BigQuery --
def query_bigquery(molecule: dict) -> list:
    if not BQ_SERVICE_ACCOUNT:
        log.err("BQ_SERVICE_ACCOUNT not set in .env -- skipping BigQuery.")
        return []
    if not Path(BQ_SERVICE_ACCOUNT).exists():
        log.err(f"Key file not found: {BQ_SERVICE_ACCOUNT}")
        return []
    try:
        from google.cloud import bigquery
        client = bigquery.Client.from_service_account_json(BQ_SERVICE_ACCOUNT)
    except ImportError:
        log.err("google-cloud-bigquery not installed."); return []
    except Exception as e:
        log.err(f"BigQuery auth failed: {e}"); return []

    terms = molecule.get("synonyms", [molecule["generic_name"].lower()])
    where = " OR ".join(f"LOWER(molecule_name) LIKE '%{t}%'" for t in terms)
    sql = f"""
        SELECT trial_id, phase, trial_size, company_name,
               hba1c_change_pct, weight_change_pct, molecule_name
        FROM `cognito-prod-394707.cognito_prod_datamart.clinical_efficacy`
        WHERE {where}
        ORDER BY phase, trial_id
    """
    log.info("Executing BigQuery query...")
    try:
        rows = []
        for row in client.query(sql).result():
            hba1c_val = None
            weight_val = None
            try:
                if row.hba1c_change_pct is not None:
                    hba1c_val = float(row.hba1c_change_pct)
            except (ValueError, TypeError):
                pass
            try:
                if row.weight_change_pct is not None:
                    weight_val = float(row.weight_change_pct)
            except (ValueError, TypeError):
                pass
            rows.append({
                "trial_id": row.trial_id, "phase": row.phase,
                "trial_size": row.trial_size, "company_name": row.company_name,
                "hba1c_change_pct": hba1c_val, "weight_change_pct": weight_val,
                "molecule_name": row.molecule_name, "source": "BigQuery",
                "results_available": (hba1c_val is not None or weight_val is not None),
            })
        log.ok(f"Retrieved {len(rows)} trials from BigQuery")
        return rows
    except Exception as e:
        log.err(f"BigQuery query error: {e}"); return []


def fetch_bq_molecules() -> list:
    """
    Returns a sorted list of distinct molecule names from the
    clinical_efficacy BigQuery table.
    Falls back to an empty list if BQ is unavailable.
    """
    if not BQ_SERVICE_ACCOUNT or not Path(BQ_SERVICE_ACCOUNT).exists():
        log.warn("BQ_SERVICE_ACCOUNT not set -- cannot fetch molecule list from BigQuery.")
        return []
    try:
        from google.cloud import bigquery
        client = bigquery.Client.from_service_account_json(BQ_SERVICE_ACCOUNT)
    except ImportError:
        log.err("google-cloud-bigquery not installed."); return []
    except Exception as e:
        log.err(f"BigQuery auth failed: {e}"); return []

    sql = """
        SELECT DISTINCT TRIM(molecule_name) AS molecule_name
        FROM `cognito-prod-394707.cognito_prod_datamart.clinical_efficacy`
        WHERE molecule_name IS NOT NULL AND TRIM(molecule_name) != ''
        ORDER BY molecule_name
    """
    try:
        results = [row.molecule_name for row in client.query(sql).result()]
        log.ok(f"Found {len(results)} distinct molecules in BigQuery table.")
        return results
    except Exception as e:
        log.err(f"Failed to fetch molecule list from BigQuery: {e}")
        return []


# -- 1B: ClinicalTrials.gov --
def classify_design(design_info: dict) -> str:
    alloc = design_info.get("allocation", "").upper()
    model = design_info.get("interventionModel", "").upper()
    stype = design_info.get("studyType", "").upper()
    if "RANDOMIZED" in alloc and "NON" not in alloc:
        return "Randomized Controlled Trial (RCT)"
    if "NON_RANDOMIZED" in alloc or "NON-RANDOMIZED" in alloc:
        return "Non-Randomized Controlled Study"
    if "SINGLE_GROUP" in model or "SINGLE GROUP" in model:
        return "Single-arm Study"
    if "OBSERVATIONAL" in stype:
        return "Observational / Real-world Study"
    return f"Other ({alloc or model or stype or 'Unknown'})"


def extract_comparators(arm_groups: list) -> list:
    comps = []
    for arm in arm_groups:
        atype = arm.get("type", "").upper()
        if atype in ("PLACEBO_COMPARATOR", "ACTIVE_COMPARATOR",
                      "SHAM_COMPARATOR", "NO_INTERVENTION"):
            comps.append({
                "arm_label": arm.get("label", "Unknown"),
                "arm_type": arm.get("type", "Unknown"),
                "description": arm.get("description", ""),
            })
    return comps or [{"arm_label": "None identified", "arm_type": "N/A", "description": ""}]


def extract_location_countries(contacts_mod: dict) -> list:
    """Return list of country names."""
    locs = contacts_mod.get("locations", [])
    countries = list(OrderedDict.fromkeys(
        l.get("country", "") for l in locs if l.get("country")
    ))
    return countries[:15]


def parse_ctgov_study(study: dict) -> dict:
    proto = study.get("protocolSection", {})
    id_mod       = proto.get("identificationModule", {})
    status_mod   = proto.get("statusModule", {})
    design_mod   = proto.get("designModule", {})
    sponsor_mod  = proto.get("sponsorCollaboratorsModule", {})
    arms_mod     = proto.get("armsInterventionsModule", {})
    contacts_mod = proto.get("contactsLocationsModule", {})

    phases = design_mod.get("phases", [])
    phase_str = ", ".join(phases) if phases else "N/A"
    design_info = design_mod.get("designInfo", {})
    design_info["studyType"] = design_mod.get("studyType", "")
    enrollment = design_mod.get("enrollmentInfo", {}).get("count", None)
    has_results = study.get("hasResults", False)
    arm_groups = arms_mod.get("armGroups", [])

    start = status_mod.get("startDateStruct", {}).get("date", "")
    end   = status_mod.get("primaryCompletionDateStruct", {}).get("date", "")
    duration = f"{start} to {end}" if start and end else "Not specified"

    countries = extract_location_countries(contacts_mod)

    return {
        "trial_id":          id_mod.get("nctId", ""),
        "title":             id_mod.get("officialTitle", id_mod.get("briefTitle", "")),
        "phase":             phase_str,
        "study_design":      classify_design(design_info),
        "trial_size":        enrollment,
        "trial_location":    ", ".join(countries) if countries else "Not specified",
        "trial_countries":   countries,
        "comparators":       extract_comparators(arm_groups),
        "duration":          duration,
        "status":            status_mod.get("overallStatus", ""),
        "company_name":      sponsor_mod.get("leadSponsor", {}).get("name", "Unknown"),
        "results_available": has_results,
        "hba1c_change_pct":  None,
        "weight_change_pct": None,
        "molecule_name":     "",
        "source":            "ClinicalTrials.gov",
    }


def _ctgov_search(query_intr: str, max_results: int = 20) -> list:
    all_studies = []
    page_token = None
    while len(all_studies) < max_results:
        params = {
            "query.intr": query_intr,
            "pageSize": min(20, max_results - len(all_studies)),
            "format": "json",
        }
        if page_token:
            params["pageToken"] = page_token
        r = http_get(CTGOV_API, params=params)
        if not r or r.status_code != 200:
            break
        data = r.json()
        studies = data.get("studies", [])
        if not studies:
            break
        all_studies.extend(studies)
        page_token = data.get("nextPageToken")
        if not page_token:
            break
        time.sleep(0.3)
    return all_studies


def global_trial_scan(molecule: dict, max_per_query: int = 20) -> list:
    generic = molecule["generic_name"]
    sponsor = molecule.get("sponsor", "")
    indications = molecule.get("indications", [])

    queries = [generic, f"{generic} phase 3", f"{generic} phase 2",
               f"{generic} randomized trial"]
    for ind in indications[:4]:
        queries.append(f"{generic} {ind}")
    if sponsor and sponsor != "Unknown":
        queries.append(f"{generic} {sponsor.split('/')[0].strip()}")

    # Combination therapies (e.g. "Cagrilintide+Semaglutide") are rarely
    # registered verbatim with that exact separator. Add the alternate
    # spellings (slash, "and", space-joined, reversed order) as additional
    # searches so trials registered under a different convention are still
    # found. We deliberately do NOT search each component in isolation
    # here -- that would pull in single-ingredient trials unrelated to the
    # combination and pollute the result set.
    if molecule.get("is_combination") and molecule.get("components"):
        for variant in _combo_name_variants(molecule["components"]):
            if variant.lower() != generic.lower():
                queries.append(variant)

    # De-duplicate queries (case-insensitive) while preserving order.
    seen_q = set()
    deduped_queries = []
    for q in queries:
        qk = q.lower().strip()
        if qk and qk not in seen_q:
            seen_q.add(qk)
            deduped_queries.append(q)
    queries = deduped_queries

    seen_nct = set()
    all_trials = []
    for q in queries:
        log.info(f"  Searching: '{q}'")
        raw = _ctgov_search(q, max_results=max_per_query)
        new = 0
        for study in raw:
            nct = (study.get("protocolSection", {})
                   .get("identificationModule", {}).get("nctId", ""))
            if nct and nct not in seen_nct:
                seen_nct.add(nct)
                rec = parse_ctgov_study(study)
                rec["molecule_name"] = generic
                all_trials.append(rec)
                new += 1
        log.info(f"    -> {new} new (total: {len(all_trials)})")
        time.sleep(0.4)
    log.ok(f"Global scan: {len(all_trials)} unique trials")
    return all_trials


def enrich_trial_from_ctgov(nct_id: str) -> dict:
    enrichment = {"study_design": "Unknown", "comparators": [], "title": "",
                  "status": "", "trial_location": "", "trial_countries": [],
                  "duration": ""}
    if not nct_id or not nct_id.upper().startswith("NCT"):
        return enrichment
    try:
        r = http_get(f"{CTGOV_API}/{nct_id}", params={"format": "json"})
        if not r or r.status_code != 200:
            return enrichment
        parsed = parse_ctgov_study(r.json())
        for k in enrichment:
            enrichment[k] = parsed.get(k, enrichment[k])
        enrichment["results_available"] = parsed.get("results_available", False)
    except Exception as e:
        log.warn(f"Enrichment failed for {nct_id}: {e}")
    return enrichment


def enrich_bq_trials(bq_rows: list) -> list:
    log.info(f"Enriching {len(bq_rows)} BQ trials...")
    for i, trial in enumerate(bq_rows):
        nct = trial.get("trial_id", "")
        log.info(f"  [{i+1}/{len(bq_rows)}] {nct}")
        enrichment = enrich_trial_from_ctgov(nct)
        trial.update(enrichment)
        if trial.get("hba1c_change_pct") is not None or trial.get("weight_change_pct") is not None:
            trial["results_available"] = True
        time.sleep(0.3)
    return bq_rows


def merge_trial_lists(*trial_lists):
    """Merge multiple trial lists, de-duplicating by trial_id. First list wins."""
    merged = OrderedDict()
    for tlist in trial_lists:
        for t in tlist:
            tid = t.get("trial_id", "")
            if tid and tid not in merged:
                merged[tid] = t
    return list(merged.values())


# -- 1D: EU Clinical Trials Register (clinicaltrialsregister.eu) --
EU_CTR_SEARCH = "https://www.clinicaltrialsregister.eu/ctr-search/search"

def search_eu_ctr(molecule: dict, max_results: int = 20) -> list:
    """Search EU Clinical Trials Register via its web search."""
    generic = molecule["generic_name"]
    log.info(f"Searching EU CTR for '{generic}'...")

    trials = []
    r = http_get(EU_CTR_SEARCH, params={"query": generic, "output": "xml"}, timeout=20)
    if not r or r.status_code != 200:
        # Try HTML fallback
        r = http_get(EU_CTR_SEARCH, params={"query": generic}, timeout=20)
        if not r or r.status_code != 200:
            log.warn("EU CTR search failed or unreachable")
            return []

    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(r.text, "html.parser")

        # Parse EudraCT result entries
        result_blocks = soup.select(".result, tr.result, div.result")
        if not result_blocks:
            # Try alternative selectors
            result_blocks = soup.find_all("table", class_="result")

        for block in result_blocks[:max_results]:
            trial = _parse_eu_ctr_block(block, generic)
            if trial and trial.get("trial_id"):
                trials.append(trial)

        # If structured parsing failed, try link-based extraction
        if not trials:
            for link in soup.find_all("a", href=True):
                href = link.get("href", "")
                if "ctr-search/trial" in href or "EudraCT" in link.get_text():
                    eudract = re.search(r"(\d{4}-\d{6}-\d{2})", href + link.get_text())
                    if eudract:
                        tid = eudract.group(1)
                        if tid and not any(t["trial_id"] == tid for t in trials):
                            title_text = link.get_text(strip=True)[:200]
                            trials.append({
                                "trial_id": tid,
                                "title": title_text,
                                "phase": "N/A",
                                "study_design": "Unknown",
                                "trial_size": None,
                                "trial_location": "Europe",
                                "trial_countries": ["European Union"],
                                "comparators": [{"arm_label": "None identified", "arm_type": "N/A", "description": ""}],
                                "duration": "Not specified",
                                "status": "Unknown",
                                "company_name": molecule.get("sponsor", "Unknown"),
                                "results_available": False,
                                "hba1c_change_pct": None,
                                "weight_change_pct": None,
                                "molecule_name": generic,
                                "source": "EU CTR",
                            })
                            if len(trials) >= max_results:
                                break

    except ImportError:
        log.warn("beautifulsoup4 needed for EU CTR parsing")
    except Exception as e:
        log.warn(f"EU CTR parse error: {e}")

    log.ok(f"EU CTR: {len(trials)} trials found")
    return trials


def _parse_eu_ctr_block(block, generic: str) -> dict:
    """Parse a single EU CTR result block."""
    text = block.get_text(" ", strip=True)
    eudract = re.search(r"(\d{4}-\d{6}-\d{2})", text)
    if not eudract:
        return {}

    # Phase detection
    phase = "N/A"
    for p_pat, p_label in [(r"Phase\s*IV", "PHASE4"), (r"Phase\s*III", "PHASE3"),
                            (r"Phase\s*II", "PHASE2"), (r"Phase\s*I\b", "PHASE1")]:
        if re.search(p_pat, text, re.IGNORECASE):
            phase = p_label; break

    return {
        "trial_id": eudract.group(1),
        "title": text[:200],
        "phase": phase,
        "study_design": "Unknown",
        "trial_size": None,
        "trial_location": "Europe",
        "trial_countries": ["European Union"],
        "comparators": [{"arm_label": "None identified", "arm_type": "N/A", "description": ""}],
        "duration": "Not specified",
        "status": "Unknown",
        "company_name": "",
        "results_available": False,
        "hba1c_change_pct": None,
        "weight_change_pct": None,
        "molecule_name": generic,
        "source": "EU CTR",
    }


# -- 1E: WHO ICTRP (trialsearch.who.int) --
WHO_ICTRP_SEARCH = "https://trialsearch.who.int/Trial2.aspx"
WHO_ICTRP_API = "https://trialsearch.who.int/api/search"

def search_who_ictrp(molecule: dict, max_results: int = 20) -> list:
    """Search WHO International Clinical Trials Registry Platform."""
    generic = molecule["generic_name"]
    log.info(f"Searching WHO ICTRP for '{generic}'...")

    trials = []

    # Try the search page
    r = http_get("https://trialsearch.who.int/Default.aspx", timeout=20)
    search_url = f"https://trialsearch.who.int/Trial2.aspx?ConditionID=&TrialID=&SearchFor={generic.replace(' ','+')}"
    r2 = http_get(search_url, timeout=20)

    if not r2 or r2.status_code != 200:
        # Fallback: try direct URL patterns
        r2 = http_get(
            "https://trialsearch.who.int/default.aspx",
            params={"SearchFor": generic}, timeout=20
        )

    if r2 and r2.status_code == 200:
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(r2.text, "html.parser")

            # WHO ICTRP returns results in table rows or div blocks
            for row in soup.find_all(["tr", "div"], class_=re.compile(r"result|trial", re.I)):
                text = row.get_text(" ", strip=True)
                trial = _parse_who_result(text, generic)
                if trial:
                    trials.append(trial)
                    if len(trials) >= max_results:
                        break

            # Alternative: look for trial ID patterns in all links
            if not trials:
                for link in soup.find_all("a", href=True):
                    href = link.get("href", "")
                    text = link.get_text(strip=True)
                    # WHO ICTRP links to various registries
                    trial_id = None
                    # ISRCTN pattern
                    m = re.search(r"(ISRCTN\d+)", text + href, re.IGNORECASE)
                    if m: trial_id = m.group(1)
                    # CTRI pattern (India)
                    m = re.search(r"(CTRI/\d{4}/\d+/\d+)", text + href)
                    if m: trial_id = m.group(1)
                    # JPRN pattern (Japan)
                    m = re.search(r"(JPRN-\S+)", text + href)
                    if m: trial_id = m.group(1)
                    # ACTRN pattern (Australia/NZ)
                    m = re.search(r"(ACTRN\d+)", text + href, re.IGNORECASE)
                    if m: trial_id = m.group(1)
                    # ChiCTR pattern (China)
                    m = re.search(r"(ChiCTR\S+)", text + href)
                    if m: trial_id = m.group(1)

                    if trial_id and not any(t["trial_id"] == trial_id for t in trials):
                        trials.append({
                            "trial_id": trial_id,
                            "title": text[:200] or trial_id,
                            "phase": "N/A",
                            "study_design": "Unknown",
                            "trial_size": None,
                            "trial_location": "International",
                            "trial_countries": [],
                            "comparators": [{"arm_label": "None identified", "arm_type": "N/A", "description": ""}],
                            "duration": "Not specified",
                            "status": "Unknown",
                            "company_name": molecule.get("sponsor", "Unknown"),
                            "results_available": False,
                            "hba1c_change_pct": None,
                            "weight_change_pct": None,
                            "molecule_name": generic,
                            "source": "WHO ICTRP",
                        })
                        if len(trials) >= max_results:
                            break

        except ImportError:
            log.warn("beautifulsoup4 needed for WHO ICTRP parsing")
        except Exception as e:
            log.warn(f"WHO ICTRP parse error: {e}")
    else:
        log.warn("WHO ICTRP unreachable")

    log.ok(f"WHO ICTRP: {len(trials)} trials found")
    return trials


def _parse_who_result(text: str, generic: str) -> dict:
    """Parse a WHO ICTRP result text block."""
    if generic.lower() not in text.lower():
        return None

    # Try to find a trial ID
    trial_id = None
    for pattern in [r"(NCT\d+)", r"(ISRCTN\d+)", r"(\d{4}-\d{6}-\d{2})",
                    r"(CTRI/\d{4}/\d+/\d+)", r"(ACTRN\d+)", r"(ChiCTR\S+)",
                    r"(JPRN-\S+)", r"(KCT\d+)", r"(DRKS\d+)"]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            trial_id = m.group(1)
            break

    if not trial_id:
        return None

    phase = "N/A"
    for p_pat, p_label in [(r"Phase\s*4", "PHASE4"), (r"Phase\s*3", "PHASE3"),
                            (r"Phase\s*2", "PHASE2"), (r"Phase\s*1", "PHASE1"),
                            (r"Phase\s*IV", "PHASE4"), (r"Phase\s*III", "PHASE3"),
                            (r"Phase\s*II", "PHASE2"), (r"Phase\s*I\b", "PHASE1")]:
        if re.search(p_pat, text, re.IGNORECASE):
            phase = p_label; break

    return {
        "trial_id": trial_id,
        "title": text[:200],
        "phase": phase,
        "study_design": "Unknown",
        "trial_size": None,
        "trial_location": "International",
        "trial_countries": [],
        "comparators": [{"arm_label": "None identified", "arm_type": "N/A", "description": ""}],
        "duration": "Not specified",
        "status": "Unknown",
        "company_name": "",
        "results_available": False,
        "hba1c_change_pct": None,
        "weight_change_pct": None,
        "molecule_name": generic,
        "source": "WHO ICTRP",
    }


# ===================================================================
# STEP 2 -- Innovator Website Deep Dive
# ===================================================================
SPONSOR_DOMAINS = {
    "novo nordisk": "novonordisk.com", "eli lilly": "lilly.com",
    "regeneron": "regeneron.com", "sanofi": "sanofi.com",
    "merck": "merck.com", "pfizer": "pfizer.com",
    "roche": "roche.com", "novartis": "novartis.com",
    "astrazeneca": "astrazeneca.com", "bristol-myers squibb": "bms.com",
    "johnson & johnson": "jnj.com", "abbvie": "abbvie.com",
    "amgen": "amgen.com", "gilead": "gilead.com", "gsk": "gsk.com",
    "bayer": "bayer.com", "takeda": "takeda.com",
    "boehringer ingelheim": "boehringer-ingelheim.com",
}

def _sponsor_domain(sponsor):
    for key, domain in SPONSOR_DOMAINS.items():
        if key in sponsor.lower(): return domain
    return sponsor.lower().replace(" ", "").replace("/", "") + ".com"


def innovator_deep_dive(molecule):
    generic = molecule["generic_name"]
    sponsor = molecule.get("sponsor", "Unknown")
    domain = _sponsor_domain(sponsor)
    sp_short = sponsor.split("/")[0].strip().replace(" ", "+")

    result = {
        "sponsor": sponsor,
        "fda_product_entries": [],
        "search_queries_suggested": [
            f"{generic} trial results site:{domain}",
            f"{generic} clinical study site:{domain}",
            f"{generic} phase 2 results site:{domain}",
            f"{generic} phase 3 results site:{domain}",
            f"{generic} investor presentation site:{domain}",
        ],
        "manual_search_links": {
            "sponsor_website":        f"https://www.google.com/search?q={generic.replace(' ','+')}+trial+results+site:{domain}",
            "press_releases":         f"https://www.google.com/search?q={generic.replace(' ','+')}+press+release+{sp_short}",
            "investor_presentations": f"https://www.google.com/search?q={generic.replace(' ','+')}+investor+presentation+{sp_short}",
            "pipeline_page":          f"https://www.google.com/search?q={sp_short}+pipeline+{generic.replace(' ','+')}",
            "clinical_trial_updates": f"https://www.google.com/search?q={generic.replace(' ','+')}+clinical+trial+update+{sp_short}",
        },
    }
    log.info(f"Checking FDA for '{generic}'...")
    fda = _search_fda(generic)
    result["fda_product_entries"] = fda
    if fda: log.ok(f"Found {len(fda)} FDA product entries")
    else:   log.info("No FDA entries (may be pre-approval)")
    return result


def _search_fda(name):
    for field in ["generic_name", "brand_name"]:
        r = http_get(FDA_API, params={"search": f'openfda.{field}:"{name}"', "limit": 10})
        if r and r.status_code == 200:
            try:
                entries = []
                for res in r.json().get("results", []):
                    openfda = res.get("openfda", {})
                    for prod in res.get("products", []):
                        entries.append({
                            "brand_name": ", ".join(openfda.get("brand_name", ["N/A"])),
                            "manufacturer": ", ".join(openfda.get("manufacturer_name", ["N/A"])),
                            "application_no": res.get("application_number", "N/A"),
                            "dosage_form": prod.get("dosage_form", "N/A"),
                            "route": prod.get("route", "N/A"),
                            "marketing_status": prod.get("marketing_status", "N/A"),
                        })
                if entries: return entries
            except Exception: pass
    return []


# ===================================================================
# STEP 3 -- Scientific Literature Search (removed)
# Publications are not fetched; publications list is always empty.
# ===================================================================


# STEP 4 -- Evidence Quality Assessment (kept for internal use)
# ===================================================================
def assess_all_evidence(trials, publications, fda_entries):
    phase_counts = {"Phase 1": 0, "Phase 2": 0, "Phase 3/4": 0, "Other": 0}
    for t in trials:
        p = str(t.get("phase", "")).upper()
        if "3" in p or "4" in p:   phase_counts["Phase 3/4"] += 1
        elif "2" in p:             phase_counts["Phase 2"] += 1
        elif "1" in p:             phase_counts["Phase 1"] += 1
        else:                      phase_counts["Other"] += 1

    design_counts = {}
    for t in trials:
        d = t.get("study_design", "Unknown")
        design_counts[d] = design_counts.get(d, 0) + 1

    rct_count = design_counts.get("Randomized Controlled Trial (RCT)", 0)
    trials_with_results = sum(1 for t in trials if t.get("results_available"))
    sizes = []
    for t in trials:
        try:
            s = int(t.get("trial_size") or 0)
            if s > 0: sizes.append(s)
        except (ValueError, TypeError): pass
    total_patients = sum(sizes)
    max_size = max(sizes) if sizes else 0

    has_fda = len(fda_entries) > 0
    has_p3 = phase_counts["Phase 3/4"] > 0

    if has_p3 and trials_with_results >= 2 and (rct_count >= 2 or has_fda):
        level, detail = "STRONG", "Multiple late-stage studies with published results"
    elif trials_with_results >= 1 and rct_count >= 1:
        level, detail = "MODERATE", "At least one controlled study with results"
    elif len(trials) > 0 and (phase_counts["Phase 1"] + phase_counts["Phase 2"]) > 0:
        level, detail = "LIMITED", "Early-stage studies only"
    else:
        level, detail = "INSUFFICIENT", "No available results"

    return {
        "evidence_level": level, "evidence_detail": detail,
        "total_trials": len(trials), "trials_with_results": trials_with_results,
        "total_publications": len(publications), "total_patients": total_patients,
        "largest_trial": max_size, "rct_count": rct_count,
        "has_regulatory_approval": has_fda,
        "phase_distribution": phase_counts, "design_distribution": design_counts,
    }


# ===================================================================
# STEP 5 -- Evidence Consolidation
# ===================================================================
def consolidation_summary(trials, publications, assessment):
    """
    Step 5: Consolidation based on TRIALS and PEER-REVIEWED PUBLICATIONS only.
    Tracks exact sources for consistency assessment.
    """
    def _to_float_pair(key):
        """Return list of (trial_id, float_value) for given endpoint key."""
        pairs = []
        for t in trials:
            raw = t.get(key)
            if raw is None: continue
            try: pairs.append((t.get("trial_id", "Unknown"), float(raw)))
            except (ValueError, TypeError): pass
        return pairs

    hba1c_pairs = _to_float_pair("hba1c_change_pct")
    weight_pairs = _to_float_pair("weight_change_pct")

    consistency = "Not assessable (insufficient endpoint data from BQ)"
    consistency_sources = []
    parts = []

    if len(hba1c_pairs) >= 2:
        ids = [p[0] for p in hba1c_pairs]
        vals = [p[1] for p in hba1c_pairs]
        if all(v < 0 for v in vals) or all(v > 0 for v in vals):
            parts.append("HbA1c: CONSISTENT")
            consistency_sources.append(f"HbA1c consistent across trials: {', '.join(ids)}")
        else:
            parts.append("HbA1c: MIXED")
            pos = [p[0] for p in hba1c_pairs if p[1] > 0]
            neg = [p[0] for p in hba1c_pairs if p[1] < 0]
            consistency_sources.append(
                f"HbA1c mixed -- positive in [{', '.join(pos)}], negative in [{', '.join(neg)}]"
            )

    if len(weight_pairs) >= 2:
        ids = [p[0] for p in weight_pairs]
        vals = [p[1] for p in weight_pairs]
        if all(v < 0 for v in vals) or all(v > 0 for v in vals):
            parts.append("Weight: CONSISTENT")
            consistency_sources.append(f"Weight consistent across trials: {', '.join(ids)}")
        else:
            parts.append("Weight: MIXED")
            pos = [p[0] for p in weight_pairs if p[1] > 0]
            neg = [p[0] for p in weight_pairs if p[1] < 0]
            consistency_sources.append(
                f"Weight mixed -- positive in [{', '.join(pos)}], negative in [{', '.join(neg)}]"
            )

    # Publication-based consistency: check if multiple publications report same phase outcomes
    pub_phase_outcomes = {}
    for p in publications:
        phase = p.get("study_phase", "")
        outcomes = p.get("key_outcomes", "")
        pmid = p.get("pmid", "") or p.get("doi", "") or "unknown"
        if phase and "Not specified" not in phase:
            if phase not in pub_phase_outcomes:
                pub_phase_outcomes[phase] = []
            pub_phase_outcomes[phase].append(pmid)

    for phase, pmids in pub_phase_outcomes.items():
        if len(pmids) >= 2:
            parts.append(f"Publications: {phase} replicated")
            consistency_sources.append(
                f"{len(pmids)} publications for {phase}: PMIDs/DOIs [{', '.join(pmids[:5])}]"
            )

    if parts:
        consistency = "; ".join(parts)

    # Replication tracking
    phase_result_counts = {}
    phase_result_ids = {}
    for t in trials:
        if t.get("results_available"):
            p = str(t.get("phase", "")).upper()
            k = "Phase 3/4" if ("3" in p or "4" in p) else ("Phase 2" if "2" in p else "Other")
            phase_result_counts[k] = phase_result_counts.get(k, 0) + 1
            phase_result_ids.setdefault(k, []).append(t.get("trial_id", ""))
    replicated = []
    for k, v in phase_result_counts.items():
        if v >= 2:
            ids = ", ".join(phase_result_ids[k][:5])
            replicated.append(f"{k} ({v} studies: {ids})")

    return {
        "total_studies": assessment["total_trials"],
        "studies_with_results": assessment["trials_with_results"],
        "results_published": len(publications) > 0,
        "peer_reviewed_count": len(publications),
        "findings_consistency": consistency,
        "consistency_sources": consistency_sources,
        "replication": ", ".join(replicated) if replicated else "No replicated phases",
        "evidence_level": assessment["evidence_level"],
        "evidence_detail": assessment["evidence_detail"],
    }


# ===================================================================
# STEP 6 -- PDGEC Scoring
# ===================================================================
TIER1_COUNTRIES = {
    "United States", "United Kingdom", "Germany", "France", "Italy", "Spain",
    "Netherlands", "Belgium", "Austria", "Sweden", "Denmark", "Finland",
    "Norway", "Ireland", "Portugal", "Poland", "Czech Republic", "Greece",
    "Hungary", "Romania", "Bulgaria", "Croatia", "Slovakia", "Slovenia",
    "Estonia", "Latvia", "Lithuania", "Luxembourg", "Malta", "Cyprus", "Iceland",
    # Common abbreviation forms from ClinicalTrials.gov
    "US", "USA", "UK",
}

TIER2_COUNTRIES = {"Canada", "Switzerland", "Australia", "Japan"}


def _get_highest_phase(trials: list) -> int:
    """Return numeric highest phase across all trials: 4,3,2,1 or 0."""
    best = 0
    for t in trials:
        p = str(t.get("phase", "")).upper()
        if "4" in p: best = max(best, 4)
        elif "3" in p: best = max(best, 3)
        elif "2" in p: best = max(best, 2)
        elif "1" in p: best = max(best, 1)
    return best


def score_phase(trials: list) -> tuple:
    """P score based on highest phase completed trial. Returns (score, reasoning)."""
    highest = _get_highest_phase(trials)
    phase_counts = {}
    for t in trials:
        p = str(t.get("phase", "")).upper()
        if "3" in p or "4" in p:   phase_counts["Phase 3/4"] = phase_counts.get("Phase 3/4", 0) + 1
        elif "2" in p:             phase_counts["Phase 2"] = phase_counts.get("Phase 2", 0) + 1
        elif "1" in p:             phase_counts["Phase 1"] = phase_counts.get("Phase 1", 0) + 1

    score = {4: 5, 3: 5, 2: 4, 1: 2}.get(highest, 0)

    if highest >= 3:
        dist = "; ".join(f"{k}: {v}" for k, v in phase_counts.items() if v > 0)
        reasoning = (f"Highest completed phase is Phase {highest}, indicating late-stage clinical development.\n"
                     f"Phase distribution: {dist}.")
    elif highest == 2:
        reasoning = (f"Highest completed phase is Phase 2, indicating mid-stage clinical development.\n"
                     f"No Phase 3 trials found; efficacy signals are preliminary.")
    elif highest == 1:
        reasoning = (f"Only Phase 1 trials identified, indicating early-stage clinical exploration.\n"
                     f"Safety and dosing data may exist but no efficacy confirmation yet.")
    else:
        reasoning = (f"No phased clinical trials identified for this molecule.\n"
                     f"The molecule may be preclinical or the trials are not yet registered.")

    return score, reasoning


def score_design(trials: list) -> tuple:
    """D score based on best study design found. Returns (score, reasoning)."""
    designs = [t.get("study_design", "") for t in trials]
    rct_count = sum(1 for d in designs if "RCT" in d or "Randomized" in d)
    nonrand_count = sum(1 for d in designs if "Non-Randomized" in d)
    single_count = sum(1 for d in designs if "Single-arm" in d)
    obs_count = sum(1 for d in designs if "Observational" in d)

    if rct_count > 0:
        score = 5
        reasoning = (f"{rct_count} randomized controlled trial(s) found, the gold standard for clinical evidence.\n"
                     f"RCT design minimizes bias and provides the strongest causal inference.")
    elif nonrand_count > 0:
        score = 3
        reasoning = (f"{nonrand_count} non-randomized controlled or multi-center study(ies) found.\n"
                     f"Controlled but without randomization; moderate risk of selection bias.")
    elif single_count > 0:
        score = 2
        reasoning = (f"{single_count} single-arm exploratory study(ies) found, with no control group.\n"
                     f"Useful for early signal detection but cannot establish comparative efficacy.")
    elif obs_count > 0:
        score = 2
        reasoning = (f"{obs_count} observational/real-world study(ies) found.\n"
                     f"Provides real-world context but is subject to confounding and selection bias.")
    else:
        score = 0
        reasoning = (f"No meaningful study design information could be identified.\n"
                     f"Trials may be registered without design details or may not exist.")

    return score, reasoning


def score_geography(trials: list) -> tuple:
    """G score based on geography tier of trial locations. Returns (score, reasoning)."""
    all_countries = set()
    for t in trials:
        countries = t.get("trial_countries", [])
        if not countries:
            loc = t.get("trial_location", "")
            if loc and loc != "Not specified":
                countries = [c.strip() for c in loc.split(",")]
        all_countries.update(countries)

    tier1_found = all_countries & TIER1_COUNTRIES
    tier2_found = all_countries & TIER2_COUNTRIES
    tier3_found = all_countries - TIER1_COUNTRIES - TIER2_COUNTRIES

    if tier1_found:
        score = 5
        t1_sample = ", ".join(sorted(tier1_found)[:5])
        reasoning = (f"Trials conducted in Tier 1 regions (US/EU/UK): {t1_sample}.\n"
                     f"These geographies have the most rigorous regulatory oversight and data quality standards.")
    elif tier2_found:
        score = 4
        t2_sample = ", ".join(sorted(tier2_found)[:5])
        reasoning = (f"Trials conducted in Tier 2 regions: {t2_sample}.\n"
                     f"Strong regulatory frameworks but not the primary markets (US/EU) for global approval.")
    elif all_countries:
        score = 3
        t3_sample = ", ".join(sorted(tier3_found)[:5]) if tier3_found else ", ".join(sorted(all_countries)[:5])
        reasoning = (f"Trials conducted in Tier 3 regions only: {t3_sample}.\n"
                     f"Results may have limited generalizability to US/EU populations and regulatory pathways.")
    else:
        score = 3
        reasoning = (f"No geographic location data available for any registered trial.\n"
                     f"Defaulting to Tier 3 score; location could not be verified from registry data.")

    return score, reasoning


def score_evidence(trials: list, fda_entries: list, rwe_synthesis: str = "") -> tuple:
    """
    E score — Evidence Support, assessed by Gemini 2.5 Flash using RWE synthesis.

    Scoring tiers:
      5  — Strong post-launch real-world evidence (approved + meaningful RWE)
      4  — FDA-approved, limited RWE or RWE still maturing
      3  — Trial results available (registry / BQ endpoints) but no post-launch RWE
      1  — Trials registered but no results yet
      0  — No clinical evidence of any kind

    Gemini reads the RWE synthesis from drug_research_summarizer and assigns the
    tier with a plain-text justification. Falls back to rule-based logic if the
    API is unavailable.
    """
    has_fda  = len(fda_entries) > 0
    trials_with_results = sum(1 for t in trials if t.get("results_available"))
    has_bq_endpoints    = any(
        t.get("hba1c_change_pct") is not None or t.get("weight_change_pct") is not None
        for t in trials
    )

    # ── Gemini path ────────────────────────────────────────────────────────
    if not _gemini_model:
        log.warn("Gemini model not initialised — using rule-based E scoring.")
    elif not rwe_synthesis:
        log.warn("RWE synthesis is empty — using rule-based E scoring.")
    else:
        prompt = f"""You are a senior pharmaceutical evidence analyst.

Your task is to assign an Evidence Support score (E) for a drug based on the
real-world evidence (RWE) synthesis below, which was generated from post-launch
real-world databases (MIMIC-IV, CMS Medicare, ResDAC, CDC, All of Us, etc.) and
the innovator's own clinical data.

SCORING RULES — assign EXACTLY ONE integer from this list:
  5 = Strong post-launch RWE: the drug is approved and multiple independent
      real-world databases show consistent evidence of efficacy and/or safety
      post-launch (e.g. HbA1c reduction, weight loss, CV benefit confirmed in
      Medicare claims, EHR cohorts, or NIH All of Us data).
  4 = FDA-approved but RWE is limited or still maturing — fewer than 2
      independent post-launch databases with quantitative outcome data, OR
      RWE data are mostly from a single source.
  3 = Clinical trial results exist in registries or sponsor-reported endpoint
      data (HbA1c change, weight change) but no meaningful post-launch
      real-world data is available yet.
  1 = Trials are registered but no results have been reported yet.
  0 = No clinical evidence of any kind found.

CONTEXT FLAGS (from registry data):
  FDA-approved product entries found: {has_fda}
  Trials with reported results: {trials_with_results}
  Sponsor-reported endpoints available: {has_bq_endpoints}

REAL-WORLD EVIDENCE SYNTHESIS (from 10 post-launch data sources):
{rwe_synthesis[:4000]}

REQUIRED OUTPUT FORMAT — two lines only, no JSON, no markdown, no preamble:
SCORE: <integer: 0, 1, 3, 4, or 5>
REASONING: <2-3 sentences explaining the score, citing specific RWE sources>"""

        try:
            with _gemini_semaphore:
                response = _gemini_model.generate_content(
                    prompt,
                    generation_config=genai.types.GenerationConfig(
                        temperature=0.1, max_output_tokens=512,
                    ),
                )
            raw = response.text.strip()
            score = None
            reasoning_text = ""
            for line in raw.splitlines():
                line = line.strip()
                if line.upper().startswith("SCORE:"):
                    try:
                        score = int(float(line.split(":", 1)[1].strip()))
                    except ValueError:
                        pass
                elif line.upper().startswith("REASONING:"):
                    reasoning_text = line.split(":", 1)[1].strip()
            if score is None:
                raise ValueError(f"Could not parse SCORE from response: {raw[:200]}")
            if score not in (0, 1, 3, 4, 5):
                score = min([0, 1, 3, 4, 5], key=lambda x: abs(x - score))
            log.ok(f"Gemini E score: {score}")
            return score, f"[LLM-assessed] {reasoning_text}"
        except Exception as e:
            log.warn(f"Gemini E scoring failed ({e}), falling back to rule-based.")

    # ── Rule-based fallback ────────────────────────────────────────────────
    if has_fda and rwe_synthesis:
        score = 5
        reasoning = ("FDA approval on record and RWE synthesis available from post-launch sources.\n"
                     "Indicates strong post-launch real-world evidence base.")
    elif has_fda:
        score = 4
        reasoning = ("FDA-approved product found but no RWE synthesis was generated.\n"
                     "Approval implies regulatory review; RWE depth could not be assessed.")
    elif trials_with_results > 0 or has_bq_endpoints:
        score = 3
        reasoning = (f"{trials_with_results} trial(s) with registry results or BQ endpoint data.\n"
                     f"Results exist but no post-launch real-world evidence available.")
    elif len(trials) > 0:
        score = 1
        reasoning = (f"{len(trials)} trial(s) registered but none report results yet.\n"
                     f"Clinical development underway; no efficacy or safety data available.")
    else:
        score = 0
        reasoning = ("No clinical evidence of any kind found.\n"
                     "Molecule may be preclinical, withdrawn, or not yet in human trials.")

    return score, reasoning



def _call_gemini(prompt: str, max_tokens: int = 1024) -> str:
    """Call Gemini 2.5 Flash API. Returns response text or empty string on failure."""
    if not GEMINI_API_KEY:
        log.warn("GEMINI_API_KEY not set in .env -- LLM consistency scoring unavailable.")
        return ""

    url = f"{GEMINI_API_URL}?key={GEMINI_API_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "maxOutputTokens": max_tokens,
            "temperature": 0.2,
        },
    }

    try:
        with _gemini_semaphore:
            r = requests.post(url, json=payload, timeout=60)
        if r.status_code == 200:
            data = r.json()
            candidates = data.get("candidates", [])
            if candidates:
                parts = candidates[0].get("content", {}).get("parts", [])
                if parts:
                    return parts[0].get("text", "")
        else:
            log.warn(f"Gemini API returned HTTP {r.status_code}: {r.text[:200]}")
    except Exception as e:
        log.warn(f"Gemini API call failed: {e}")

    return ""


def score_consistency(trials: list, rwe_synthesis: str, consolidation: dict) -> tuple:
    """
    C score — Consistency Adjustment, assessed by Gemini 2.5 Flash.

    Scoring thresholds:
      +0.5  — >= 20% of studies show consistent positive findings
      +0.25 — >= 10% but < 20% of studies are strong with no replication
      -0.5  — Mixed or conflicting evidence across studies
      -1.0  — Results unavailable (no results reported in any trial)

    Gemini reads trial data + the RWE synthesis and assigns a score.
    Falls back to rule-based logic if the API is unavailable.
    """
    molecule_name = next(
        (t["molecule_name"] for t in trials if t.get("molecule_name")), "Unknown"
    )

    n_trials      = len(trials)
    n_results     = sum(1 for t in trials if t.get("results_available"))
    n_rct         = sum(1 for t in trials if "randomized" in str(t.get("study_design", "")).lower())
    cons_text     = consolidation.get("findings_consistency", "Not assessed")
    replication   = consolidation.get("replication", "")
    hba1c_vals    = [t["hba1c_change_pct"] for t in trials if t.get("hba1c_change_pct") is not None]
    weight_vals   = [t["weight_change_pct"] for t in trials if t.get("weight_change_pct") is not None]
    terminated    = [t.get("trial_id", "?") for t in trials
                     if "terminat" in str(t.get("status", "")).lower()]

    # ── Gemini path ────────────────────────────────────────────────────────
    if not _gemini_model:
        log.warn("Gemini model not initialised — using rule-based C scoring.")
        return _rule_based_consistency(trials, consolidation)

    if True:  # Gemini model available — proceed
        prompt = f"""You are a senior pharmaceutical evidence analyst.

Assign a Consistency Adjustment score (C) for the molecule "{molecule_name}".

SCORING RULES — assign EXACTLY ONE value from this list:
  +0.5  = >= 20% of all registered studies show consistent POSITIVE findings
          (e.g. HbA1c reduction, weight loss confirmed in multiple independent
          trials AND corroborated by real-world data sources)
  +0.25 = >= 10% but < 20% of studies are strong/positive with no replication
          (a single high-quality study with clear positive signal, not yet
          independently replicated)
  -0.5  = Mixed or conflicting evidence: some studies positive, others negative
          or null; OR real-world data contradicts trial results
  -1.0  = Results unavailable: no trial has reported any results yet, or data
          is entirely absent from registry and real-world sources

CRITICAL RULES:
- Base the percentage on TOTAL registered trials (n={n_trials}), not just those
  with results.
- Routine adverse events (nausea, GI events, injection-site reactions) are NOT
  evidence of inconsistency. Do NOT use them to justify -0.5 or -1.0.
- -1.0 means truly NO results available, not just limited results.
- +0.25 is the appropriate score when there is exactly 1 strong positive study
  and no contradicting evidence.

TRIAL EVIDENCE:
- Total trials: {n_trials}  |  With results: {n_results}  |  RCTs: {n_rct}
- Rule-based consistency flag: {cons_text}
- Replication summary: {replication}
- HbA1c values across trials: {hba1c_vals[:20] if hba1c_vals else 'Not available'}
- Weight change values across trials: {weight_vals[:20] if weight_vals else 'Not available'}
- Terminated trials: {terminated if terminated else 'None'}

REAL-WORLD EVIDENCE SYNTHESIS (post-launch databases):
{rwe_synthesis[:3000] if rwe_synthesis else 'Not available'}

REQUIRED OUTPUT FORMAT — three lines only, no JSON, no markdown, no preamble:
SCORE: <number: one of 0.5, 0.25, -0.5, -1.0>
PCT: <float: estimated % of total trials with consistent positive findings>
REASONING: <3-4 sentences citing specific trials, endpoints, and RWE sources that drove the score>"""

        try:
            with _gemini_semaphore:
                response = _gemini_model.generate_content(
                    prompt,
                    generation_config=genai.types.GenerationConfig(
                        temperature=0.1, max_output_tokens=1024,
                    ),
                )
            raw = response.text.strip()
            score = None
            pct = 0.0
            rationale = ""
            for line in raw.splitlines():
                line = line.strip()
                if line.upper().startswith("SCORE:"):
                    try:
                        score = float(line.split(":", 1)[1].strip())
                    except ValueError:
                        pass
                elif line.upper().startswith("PCT:"):
                    try:
                        pct = float(line.split(":", 1)[1].strip().rstrip("%"))
                    except ValueError:
                        pass
                elif line.upper().startswith("REASONING:"):
                    rationale = line.split(":", 1)[1].strip()
            if score is None:
                raise ValueError(f"Could not parse SCORE from response: {raw[:200]}")
            valid_scores = [0.5, 0.25, -0.5, -1.0]
            if score not in valid_scores:
                score = min(valid_scores, key=lambda x: abs(x - score))
            reasoning_text = (
                f"[LLM-assessed] ~{pct:.1f}% of {n_trials} studies consistent positive.\n"
                f"{rationale}"
            )
            score_consistency._last_llm_detail = {"summary": rationale, "score": score, "pct": pct}
            log.ok(f"Gemini C score: {score}  ({pct:.1f}% consistent)")
            return score, reasoning_text
        except Exception as e:
            log.warn(f"Gemini C scoring failed ({e}), falling back to rule-based.")

    # ── Rule-based fallback ────────────────────────────────────────────────
    return _rule_based_consistency(trials, consolidation)


def _rule_based_consistency(trials: list, consolidation: dict) -> tuple:
    """Fallback rule-based consistency scoring using the new C thresholds."""
    n_total  = len(trials)
    n_results = sum(1 for t in trials if t.get("results_available"))
    cons     = consolidation.get("findings_consistency", "")
    result_trials = [t for t in trials if t.get("results_available")]

    # No results at all → -1.0
    if n_results == 0:
        return -1.0, (
            f"[Rule-based] No trial results available across {n_total} registered studies.\n"
            f"Consistency cannot be assessed; score is -1.0 (results unavailable)."
        )

    # Mixed signals → -0.5
    if "MIXED" in cons.upper():
        return -0.5, (
            f"[Rule-based] Mixed/conflicting endpoint signals detected across trials.\n"
            f"Consistency flag: {cons}"
        )

    # Calculate % of total trials with consistent positive results
    pct = (n_results / n_total * 100) if n_total > 0 else 0

    if pct >= 20 and "CONSISTENT" in cons.upper():
        ids = ", ".join(t.get("trial_id", "?") for t in result_trials[:5])
        return 0.5, (
            f"[Rule-based] {pct:.1f}% of {n_total} studies ({n_results}) show consistent positive findings.\n"
            f"Threshold >= 20% met. Trials: [{ids}]"
        )
    elif pct >= 10:
        ids = ", ".join(t.get("trial_id", "?") for t in result_trials[:3])
        return 0.25, (
            f"[Rule-based] {pct:.1f}% of {n_total} studies ({n_results}) are strong/positive, no replication yet.\n"
            f"Threshold >= 10% < 20% met. Trials: [{ids}]"
        )
    else:
        # Results exist but below 10% — treat as insufficient for positive adjustment
        return -0.5, (
            f"[Rule-based] Only {pct:.1f}% of {n_total} studies ({n_results}) report results.\n"
            f"Insufficient consistent evidence; treating as mixed/uncertain."
        )


def compute_final_score(P: float, D: float, G: float, E: float, C: float) -> float:
    """
    Final Score = min(5, max(1, ((P + mode(D, G, E)) / 2) + C))

    P            — Phase score (development maturity)
    mode(D, G, E) — Modal score across Study Design (D), Geography (G),
                    and Evidence Support (E), representing the overall
                    quality of the evidence body. If all three are distinct
                    (no mode), the median is used as a fallback.
    C            — Consistency adjustment
    """
    try:
        dge_mode = _stat_mode([D, G, E])
    except Exception:
        # No unique mode — use median as fallback
        dge_mode = sorted([D, G, E])[1]
    raw = ((P + dge_mode) / 2.0) + C
    return round(min(5.0, max(1.0, raw)), 2)


def identify_evidence_gaps(trials, publications, assessment, consolidation) -> list:
    """Identify gaps like a medical analyst would."""
    gaps = []
    if assessment.get("rct_count", 0) == 0:
        gaps.append("No randomized controlled trials identified")
    if assessment.get("total_publications", 0) == 0:
        gaps.append("No peer-reviewed publications found")
    if not assessment.get("has_regulatory_approval"):
        gaps.append("No FDA regulatory approval on record")

    # Long-term data
    has_long = False
    for p in publications:
        dur = (p.get("duration", "") or "").lower()
        if "year" in dur:
            try:
                yrs = int(re.search(r"(\d+)", dur).group(1))
                if yrs >= 2: has_long = True
            except Exception: pass
    if not has_long:
        gaps.append("Limited long-term safety data (no studies >= 2 years found)")

    # Head-to-head
    comps = set()
    for t in trials:
        for c in t.get("comparators", []):
            if "ACTIVE" in c.get("arm_type", "").upper():
                comps.add(c.get("arm_label", ""))
    if not comps:
        gaps.append("No head-to-head comparisons with active therapies")

    # Geographic diversity
    all_countries = set()
    for t in trials:
        cc = t.get("trial_countries", [])
        all_countries.update(cc)
    if len(all_countries) <= 2:
        gaps.append(f"Limited geographic diversity (only {len(all_countries)} countries)")

    # Replication
    if "No replicated" in consolidation.get("replication", ""):
        gaps.append("No replication of results across independent studies")

    return gaps if gaps else ["No major evidence gaps identified"]


# ===================================================================
# BQ-READY OUTPUT: Flat Tables
# ===================================================================

# -- Table 1: molecule_scoring (1 row per molecule) --
SCORING_COLUMNS = [
    "molecule_name", "brand_names", "sponsor", "dev_codes",
    "total_clinical_studies", "studies_with_results", "peer_reviewed_publications",
    "highest_trial_phase", "total_patients_enrolled", "largest_trial_size",
    "rct_count", "has_regulatory_approval",
    "score_phase_P", "reasoning_phase_P",
    "score_design_D", "reasoning_design_D",
    "score_geography_G", "reasoning_geography_G",
    "score_evidence_E", "reasoning_evidence_E",
    "score_consistency_C", "reasoning_consistency_C",
    "final_score",
    "evidence_level", "evidence_detail",
    "findings_consistency", "replication",
    "evidence_gaps",
    "report_generated_at",
]

# -- Table 2: trial_details (1 row per trial) --
TRIAL_COLUMNS = [
    "molecule_name", "trial_id", "title", "phase", "study_design",
    "trial_size", "trial_location", "comparator_arms", "duration",
    "status", "sponsor", "results_available",
    "hba1c_change_pct", "weight_change_pct",
    "source", "report_generated_at",
]

# -- Table 3: publication_details (1 row per publication) --
PUB_COLUMNS = [
    "molecule_name", "pmid", "doi", "title", "journal", "year",
    "study_phase", "sample_size", "population", "duration",
    "key_outcomes", "safety_findings", "citation_count",
    "report_generated_at",
]


def build_scoring_row(molecule, trials, publications, assessment,
                      consolidation, P, D, G, E, C, final, gaps, reasoning, ts):
    return {
        "molecule_name":           molecule["generic_name"],
        "brand_names":             "; ".join(molecule.get("brand_names", [])),
        "sponsor":                 molecule.get("sponsor", ""),
        "dev_codes":               "; ".join(molecule.get("dev_codes", [])),
        "total_clinical_studies":  assessment["total_trials"],
        "studies_with_results":    assessment["trials_with_results"],
        "peer_reviewed_publications": assessment["total_publications"],
        "highest_trial_phase":     f"Phase {_get_highest_phase(trials)}" if _get_highest_phase(trials) > 0 else "None",
        "total_patients_enrolled": assessment["total_patients"],
        "largest_trial_size":      assessment["largest_trial"],
        "rct_count":               assessment["rct_count"],
        "has_regulatory_approval": assessment["has_regulatory_approval"],
        "score_phase_P":           P,
        "reasoning_phase_P":       safe_str(reasoning.get("P", "")),
        "score_design_D":          D,
        "reasoning_design_D":      safe_str(reasoning.get("D", "")),
        "score_geography_G":       G,
        "reasoning_geography_G":   safe_str(reasoning.get("G", "")),
        "score_evidence_E":        E,
        "reasoning_evidence_E":    safe_str(reasoning.get("E", "")),
        "score_consistency_C":     C,
        "reasoning_consistency_C": safe_str(reasoning.get("C", "")),
        "final_score":             final,
        "evidence_level":          assessment["evidence_level"],
        "evidence_detail":         assessment["evidence_detail"],
        "findings_consistency":    consolidation["findings_consistency"],
        "replication":             consolidation["replication"],
        "evidence_gaps":           "; ".join(gaps),
        "report_generated_at":     ts,
    }


def build_trial_rows(molecule, trials, ts):
    rows = []
    for t in trials:
        comps = t.get("comparators", [])
        comp_str = "; ".join(
            f"{c['arm_label']} ({c['arm_type']})" for c in comps
            if c.get("arm_label") != "None identified"
        ) or "None"
        rows.append({
            "molecule_name":     molecule["generic_name"],
            "trial_id":          t.get("trial_id", ""),
            "title":             safe_str(t.get("title", "")),
            "phase":             t.get("phase", ""),
            "study_design":      t.get("study_design", ""),
            "trial_size":        t.get("trial_size", ""),
            "trial_location":    t.get("trial_location", ""),
            "comparator_arms":   comp_str,
            "duration":          t.get("duration", ""),
            "status":            t.get("status", ""),
            "sponsor":           t.get("company_name", ""),
            "results_available": t.get("results_available", False),
            "hba1c_change_pct":  t.get("hba1c_change_pct", ""),
            "weight_change_pct": t.get("weight_change_pct", ""),
            "source":            t.get("source", ""),
            "report_generated_at": ts,
        })
    return rows


def build_pub_rows(molecule, publications, ts):
    rows = []
    for p in publications:
        rows.append({
            "molecule_name":     molecule["generic_name"],
            "pmid":              p.get("pmid", ""),
            "doi":               p.get("doi", ""),
            "title":             safe_str(p.get("title", "")),
            "journal":           safe_str(p.get("journal", "")),
            "year":              p.get("year", ""),
            "study_phase":       p.get("study_phase", ""),
            "sample_size":       p.get("sample_size", ""),
            "population":        p.get("population", ""),
            "duration":          p.get("duration", ""),
            "key_outcomes":      safe_str(p.get("key_outcomes", "")),
            "safety_findings":   safe_str(p.get("safety_findings", "")),
            "citation_count":    p.get("citation_count", 0),
            "report_generated_at": ts,
        })
    return rows


def write_csv(path, columns, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_bq_to_bigquery(scoring_row, trial_rows, pub_rows):
    """Optionally write results back to BigQuery if credentials available."""
    if not BQ_SERVICE_ACCOUNT or not Path(BQ_SERVICE_ACCOUNT).exists():
        return False
    try:
        from google.cloud import bigquery
        client = bigquery.Client.from_service_account_json(BQ_SERVICE_ACCOUNT)

        project = "cognito-prod-394707"
        dataset = "cognito_prod_datamart"

        # molecule_scoring table
        table_scoring = f"{project}.{dataset}.molecule_evidence_scoring"
        errors = client.insert_rows_json(table_scoring, [scoring_row])
        if errors:
            log.warn(f"BQ insert errors (scoring): {errors}")
        else:
            log.ok(f"Inserted scoring row into {table_scoring}")

        # trial_details table
        if trial_rows:
            table_trials = f"{project}.{dataset}.molecule_trial_details"
            errors = client.insert_rows_json(table_trials, trial_rows)
            if errors:
                log.warn(f"BQ insert errors (trials): {errors}")
            else:
                log.ok(f"Inserted {len(trial_rows)} rows into {table_trials}")

        # publication_details table
        if pub_rows:
            table_pubs = f"{project}.{dataset}.molecule_publication_details"
            errors = client.insert_rows_json(table_pubs, pub_rows)
            if errors:
                log.warn(f"BQ insert errors (pubs): {errors}")
            else:
                log.ok(f"Inserted {len(pub_rows)} rows into {table_pubs}")

        return True
    except Exception as e:
        log.warn(f"BQ write-back failed: {e}")
        return False


# ===================================================================
# Report Generator
# ===================================================================
def generate_report(molecule, trials, publications, sponsor_info,
                    assessment, consolidation, P, D, G, E, C, final, gaps, reasoning=None):
    L = []
    def sec(t): L.append(f"\n{'='*72}\n  {t}\n{'='*72}")
    def sub(t): L.append(f"\n  -- {t} --")

    if reasoning is None:
        reasoning = {}

    L.append("=" * 72)
    L.append("  CLINICAL EVIDENCE SUMMARY REPORT")
    L.append(f"  Molecule : {molecule['generic_name']}")
    L.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    L.append("=" * 72)

    # Molecule
    sec("MOLECULE INFORMATION")
    L.append(f"  Generic Name  : {molecule['generic_name']}")
    L.append(f"  Brand Name(s) : {', '.join(molecule.get('brand_names', [])) or 'N/A'}")
    L.append(f"  Dev Code(s)   : {', '.join(molecule.get('dev_codes', [])) or 'N/A'}")
    L.append(f"  Sponsor       : {molecule.get('sponsor', 'Unknown')}")
    L.append(f"  Indications   : {', '.join(molecule.get('indications', [])[:8]) or 'N/A'}")

    # SCORING (Step 6)
    sec("STEP 6: PDGEC SCORING")

    L.append(f"\n  Score (P) Phase       : {P} / 5")
    if reasoning.get("P"):
        for line in reasoning["P"].split("\n"):
            L.append(f"    >> {line.strip()}")

    L.append(f"\n  Score (D) Design      : {D} / 5")
    if reasoning.get("D"):
        for line in reasoning["D"].split("\n"):
            L.append(f"    >> {line.strip()}")

    L.append(f"\n  Score (G) Geography   : {G} / 5")
    if reasoning.get("G"):
        for line in reasoning["G"].split("\n"):
            L.append(f"    >> {line.strip()}")

    L.append(f"\n  Score (E) Evidence    : {E} / 5")
    if reasoning.get("E"):
        for line in reasoning["E"].split("\n"):
            L.append(f"    >> {line.strip()}")

    L.append(f"\n  Score (C) Consistency : {C}")
    if reasoning.get("C"):
        c_text = reasoning["C"]
        if "[LLM-assessed]" in c_text:
            import textwrap as _tw
            L.append("")
            L.append("    [LLM Consistency Assessment]")
            L.append("")
            # Strip the tag and print each line numbered and wrapped
            body = c_text.replace("[LLM-assessed]", "").strip()
            for _line in body.splitlines():
                _line = _line.strip()
                if not _line:
                    continue
                for _wl in _tw.wrap(_line, width=84):
                    L.append(f"      {_wl}")
            L.append("")
        else:
            # Rule-based: render as a detailed formatted block
            L.append("")
            L.append(f"    [Rule-based assessment]")
            L.append("")
            lines = [l.strip() for l in c_text.split("\n") if l.strip()]
            # First line is typically the main finding; rest are supporting detail
            if lines:
                wrapped = _wrap_text(lines[0], width=88, indent="    ")
                L.append(wrapped)
            for line in lines[1:]:
                wrapped = _wrap_text(line, width=88, indent="    ")
                L.append(wrapped)

    L.append(f"\n  ----------------------------------------")
    L.append(f"  FINAL SCORE           : {final} / 5.0")
    L.append(f"  Formula: min(5, max(1, avg(P,D,G,E) + C))")

    # Evidence overview
    sec("EVIDENCE OVERVIEW")
    L.append(f"  Number of human clinical studies  : {assessment['total_trials']}")
    L.append(f"  Studies with available results    : {assessment['trials_with_results']}")
    L.append(f"  Peer-reviewed publications        : {assessment['total_publications']}")
    L.append(f"  Highest trial phase               : Phase {_get_highest_phase(trials)}" if _get_highest_phase(trials) > 0 else "  Highest trial phase               : None")
    L.append(f"  Total patients enrolled           : {assessment['total_patients']:,}")
    L.append(f"  Largest trial                     : {assessment['largest_trial']:,} patients")
    L.append(f"  RCT count                         : {assessment['rct_count']}")
    L.append(f"  Regulatory approval (FDA)         : {'Yes' if assessment['has_regulatory_approval'] else 'No'}")

    # Evidence strength
    sec("EVIDENCE STRENGTH")
    L.append(f"  Level   : {assessment['evidence_level']}")
    L.append(f"  Reason  : {assessment['evidence_detail']}")
    L.append(f"  Consistency : {consolidation['findings_consistency']}")
    L.append(f"  Replication : {consolidation['replication']}")

    # Evidence gaps
    sec("EVIDENCE GAPS")
    for g in gaps:
        L.append(f"  - {g}")

    # Phase & Design distribution
    sec("DISTRIBUTIONS")
    sub("Phase Distribution")
    for ph, ct in assessment["phase_distribution"].items():
        L.append(f"    {ph:<15s} : {ct}")
    sub("Study Design Distribution")
    for d, ct in assessment["design_distribution"].items():
        L.append(f"    {d:<50s} : {ct}")

    # Trials - summary counts only; individual trial detail is in the CSV/JSON outputs
    sec("CLINICAL TRIALS DETAIL")
    L.append(f"  Individual trial records are available in the companion CSV and JSON outputs.")
    L.append(f"  Aggregate statistics are reported in the EVIDENCE OVERVIEW section above.")

    # Sponsor
    sec("INNOVATOR / SPONSOR")
    L.append(f"  Sponsor: {sponsor_info.get('sponsor', 'N/A')}")
    fda = sponsor_info.get("fda_product_entries", [])
    if fda:
        sub("FDA Product Entries")
        for a in fda[:10]:
            L.append(f"    {a['brand_name']} | {a['manufacturer']} | {a['application_no']} | {a['marketing_status']}")
    sub("Suggested Searches")
    for q in sponsor_info.get("search_queries_suggested", [])[:5]:
        L.append(f"    - {q}")

    # Publications - summary counts only; individual publication detail is in the CSV/JSON outputs
    sec("SCIENTIFIC PUBLICATIONS")
    L.append(f"  Individual publication records are available in the companion CSV and JSON outputs.")
    L.append(f"  Aggregate statistics are reported in the EVIDENCE OVERVIEW section above.")

    # Regulatory
    sec("REGULATORY LINKS")
    g = molecule["generic_name"]
    L.append(f"  FDA: https://www.accessdata.fda.gov/scripts/cder/ob/results_product.cfm?Generic_Name={g}")
    L.append(f"  EMA: https://www.ema.europa.eu/en/search?search_api_fulltext={g}")

    L.append(f"\n{'='*72}\n  END OF REPORT\n{'='*72}\n")
    return "\n".join(L)


# ===================================================================
# STEP 3 (NEW) -- Real-World Evidence Summarizer
# Uses drug_research_summarizer sources + Gemini to generate a
# post-launch RWE synthesis that feeds into E and C scoring.
# ===================================================================

# RWE source prompts (condensed from drug_research_summarizer)
_RWE_SOURCES = [
    ("mimic",    "MIMIC-IV (ICU EHR, Beth Israel / MIT)"),
    ("cms",      "CMS Medicare Public Use Files"),
    ("resdac",   "ResDAC — Medicare & Medicaid longitudinal claims"),
    ("cdc",      "CDC Open Data / NHANES population surveys"),
    ("allofus",  "NIH All of Us Research Program (700k+ participants)"),
    ("innovator","Innovator / Manufacturer Website (press releases, trials, pipeline)"),
]

_RWE_PROMPT_TEMPLATE = """You are an expert clinical data analyst specialising in real-world evidence research.

Drug under investigation: "{drug}"
Data source: {source_name}

Write a concise real-world evidence summary (150-200 words) for "{drug}" from the perspective
of {source_name}. Focus ONLY on quantitative post-launch evidence:
- HbA1c reduction observed in this data source (mean Δ or % reaching <7.0%)
- Weight loss (%TBWL or kg) documented in this data source
- Cardiovascular or renal outcomes if available
- Any safety signals unique to real-world use
- Population size / cohort characteristics

If this specific source has limited or no data on "{drug}", state that clearly in 2-3 sentences.
Report numbers where known; do not fabricate specific statistics.
Keep the summary to 150-200 words."""


def _run_rwe_summarizer(drug_name: str) -> str:
    """
    Step 3: Query Gemini for post-launch RWE summaries across all RWE sources,
    then generate an integrated synthesis used for E and C scoring.
    Returns the synthesis text (or empty string if Gemini unavailable).

    Note: the per-source fetches below run in their own ThreadPoolExecutor.
    They're submitted via submit_with_ctx() so that this molecule's log
    buffer (set higher up in run_molecule) correctly follows into these
    worker threads instead of falling back to unbuffered printing. Each
    actual Gemini call is separately gated by the global _gemini_semaphore
    so a burst of 6 source calls (times however many molecules are running
    in parallel) doesn't overwhelm the API.
    """
    if not _gemini_model:
        log.warn("Gemini model unavailable — RWE synthesis skipped.")
        return ""

    log.sub("Collecting RWE summaries from data sources (parallel)...")
    source_summaries = []

    def _fetch_rwe_source(src_id, src_name):
        prompt = _RWE_PROMPT_TEMPLATE.format(drug=drug_name, source_name=src_name)
        try:
            with _gemini_semaphore:
                response = _gemini_model.generate_content(
                    prompt,
                    generation_config=genai.types.GenerationConfig(
                        temperature=0.3, max_output_tokens=512,
                    ),
                )
            summary = response.text.strip()
            log.ok(f"  {src_name}: {len(summary)} chars")
            return f"=== {src_name} ===\n{summary}"
        except Exception as e:
            log.warn(f"  {src_name} summary failed: {e}")
            return f"=== {src_name} ===\nData unavailable."

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            submit_with_ctx(executor, _fetch_rwe_source, src_id, src_name): idx
            for idx, (src_id, src_name) in enumerate(_RWE_SOURCES)
        }
        indexed_results = {}
        for future in as_completed(futures):
            idx = futures[future]
            indexed_results[idx] = future.result()
    # Preserve original order
    source_summaries = [indexed_results[i] for i in range(len(_RWE_SOURCES))]

    combined = "\n\n".join(source_summaries)

    # Integrated synthesis
    synthesis_prompt = f"""You are a senior pharmaceutical researcher.

Based on the real-world evidence summaries below from {len(source_summaries)} data sources,
write an integrated RWE synthesis (300-400 words) for "{drug_name}".

Structure your synthesis around:
1. CONSISTENCY OF FINDINGS — do multiple independent sources confirm efficacy?
   Estimate what percentage of data sources report positive/consistent findings.
2. QUANTITATIVE EVIDENCE STRENGTH — cite the strongest numeric outcomes seen
   (HbA1c Δ, %TBWL, MACE HR) across sources.
3. POST-LAUNCH REAL-WORLD PERFORMANCE — does RWE align with or diverge from
   clinical trial results?
4. EVIDENCE GAPS — what outcomes or populations lack real-world data?

--- SUMMARIES FROM {len(source_summaries)} SOURCES ---
{combined}

Write in professional clinical language. Be specific about which sources provide
strong evidence vs. which are limited or absent."""

    try:
        with _gemini_semaphore:
            response = _gemini_model.generate_content(
                synthesis_prompt,
                generation_config=genai.types.GenerationConfig(
                    temperature=0.2, max_output_tokens=2048,
                ),
            )
        synthesis = response.text.strip()
        log.ok(f"RWE synthesis generated ({len(synthesis)} chars)")
        return synthesis
    except Exception as e:
        log.warn(f"RWE synthesis failed: {e}")
        return combined   # return raw summaries as fallback


# ===================================================================
# MAIN
# ===================================================================
def run_molecule(args_molecule, args_max_trials, args_write_bq) -> dict:
    """
    Run the full pipeline for a single molecule. Returns a dict of PDGEC scores.

    All log.* calls made in here (and in every function this calls,
    including nested thread pools submitted via submit_with_ctx) write to
    whatever buffer is currently set in _log_buffer_var for this thread's
    context. When called directly from main() for a single molecule with
    no buffer set, logs print immediately as before.
    """
    log.header(f"CLINICAL EVIDENCE ANALYZER v3 -- {args_molecule.upper()}")
    t0 = time.time()

    # Step 0
    log.step(0, "MOLECULE NORMALIZATION")
    molecule = normalize_molecule(args_molecule)
    log.ok(f"Generic : {molecule['generic_name']}")
    log.info(f"Brands  : {', '.join(molecule.get('brand_names', [])) or 'N/A'}")
    log.info(f"Sponsor : {molecule.get('sponsor', 'Unknown')}")

    # Step 1
    log.step(1, "GLOBAL CLINICAL TRIAL EVIDENCE SCAN")
    log.info("Running trial source searches in parallel...")

    bq_trials = []
    ctgov_trials = []
    eu_trials = []
    who_trials = []

    def _fetch_bq():
        log.sub("1A: BigQuery")
        return query_bigquery(molecule)

    def _fetch_ctgov():
        log.sub("1B: ClinicalTrials.gov")
        return global_trial_scan(molecule, max_per_query=args_max_trials)

    def _fetch_eu():
        log.sub("1C: EU Clinical Trials Register")
        return search_eu_ctr(molecule, max_results=args_max_trials)

    def _fetch_who():
        log.sub("1D: WHO ICTRP")
        return search_who_ictrp(molecule, max_results=args_max_trials)

    source_fns = {
        "bq": _fetch_bq,
        "ctgov": _fetch_ctgov,
        "eu": _fetch_eu,
        "who": _fetch_who,
    }

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {submit_with_ctx(executor, fn): name for name, fn in source_fns.items()}
        results_map = {}
        for future in as_completed(futures):
            name = futures[future]
            try:
                results_map[name] = future.result()
            except Exception as e:
                log.warn(f"Trial source '{name}' failed: {e}")
                results_map[name] = []

    bq_trials = results_map.get("bq", [])
    ctgov_trials = results_map.get("ctgov", [])
    eu_trials = results_map.get("eu", [])
    who_trials = results_map.get("who", [])

    if bq_trials:
        log.sub("1E: Enriching BQ trials")
        bq_trials = enrich_bq_trials(bq_trials)
    trials = merge_trial_lists(bq_trials, ctgov_trials, eu_trials, who_trials)
    log.ok(f"Total unique trials: {len(trials)}")
    log.info(f"  BQ: {len(bq_trials)} | CTgov: {len(ctgov_trials)} | EU CTR: {len(eu_trials)} | WHO: {len(who_trials)}")
    log.info(f"  With results: {sum(1 for t in trials if t.get('results_available'))}")

    # Step 2
    log.step(2, "INNOVATOR WEBSITE DEEP DIVE")
    sponsor_info = innovator_deep_dive(molecule)

    # Step 3 — Real-World Evidence Summarizer
    log.step(3, "REAL-WORLD EVIDENCE SYNTHESIS")
    rwe_synthesis = _run_rwe_summarizer(molecule["generic_name"])
    publications = []   # journal fetching disabled; RWE synthesis used instead

    # Step 4
    log.step(4, "EVIDENCE QUALITY ASSESSMENT")
    assessment = assess_all_evidence(trials, publications, sponsor_info.get("fda_product_entries", []))
    log.ok(f"Evidence Level: {assessment['evidence_level']}")

    # Step 5
    log.step(5, "EVIDENCE CONSOLIDATION")
    consolidation = consolidation_summary(trials, publications, assessment)
    log.ok(f"Studies: {consolidation['total_studies']} | Results: {consolidation['studies_with_results']}")
    log.ok(f"Consistency: {consolidation['findings_consistency']}")

    # Step 6
    log.step(6, "PDGEC SCORING")
    score_consistency._last_llm_detail = None  # reset before each run
    P, P_reason = score_phase(trials)
    D, D_reason = score_design(trials)
    G, G_reason = score_geography(trials)
    E, E_reason = score_evidence(trials, sponsor_info.get("fda_product_entries", []), rwe_synthesis)
    C, C_reason = score_consistency(trials, rwe_synthesis, consolidation)
    final = compute_final_score(P, D, G, E, C)
    gaps = identify_evidence_gaps(trials, publications, assessment, consolidation)
    reasoning = {"P": P_reason, "D": D_reason, "G": G_reason, "E": E_reason, "C": C_reason}

    log.ok(f"P={P}  D={D}  G={G}  E={E}  C={C}")
    log.ok(f"FINAL SCORE = {final} / 5.0")
    log.info(f"Evidence Gaps: {'; '.join(gaps)}")

    # -- Generate outputs --
    log.header("GENERATING OUTPUTS")
    safe = re.sub(r"[^a-zA-Z0-9]", "_", molecule["generic_name"].lower())
    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    ts_bq = datetime.now().isoformat()

    report = generate_report(molecule, trials, publications, sponsor_info,
                             assessment, consolidation, P, D, G, E, C, final, gaps, reasoning)
    rpt_path = OUTPUT_DIR / f"{safe}_evidence_report_{ts_file}.txt"
    rpt_path.write_text(report, encoding="utf-8")
    log.ok(f"Report  -> {rpt_path}")

    json_path = OUTPUT_DIR / f"{safe}_evidence_data_{ts_file}.json"
    llm_consistency_detail = getattr(score_consistency, '_last_llm_detail', None)
    payload = {
        "report_generated": ts_bq, "molecule": molecule,
        "scoring": {"P": P, "D": D, "G": G, "E": E, "C": C, "final": final,
                    "reasoning": reasoning,
                    "consistency_llm_detail": llm_consistency_detail},
        "evidence_gaps": gaps,
        "evidence_consolidation": consolidation, "evidence_assessment": assessment,
        "clinical_trials": trials, "publications": publications,
        "sponsor_info": sponsor_info,
        "rwe_synthesis": rwe_synthesis,
    }
    json_path.write_text(json.dumps(payload, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
    log.ok(f"JSON    -> {json_path}")

    scoring_row = build_scoring_row(molecule, trials, publications, assessment,
                                     consolidation, P, D, G, E, C, final, gaps, reasoning, ts_bq)
    trial_rows  = build_trial_rows(molecule, trials, ts_bq)
    pub_rows    = build_pub_rows(molecule, publications, ts_bq)

    csv_scoring = OUTPUT_DIR / f"{safe}_scoring_{ts_file}.csv"
    csv_trials  = OUTPUT_DIR / f"{safe}_trials_{ts_file}.csv"
    csv_pubs    = OUTPUT_DIR / f"{safe}_publications_{ts_file}.csv"

    write_csv(csv_scoring, SCORING_COLUMNS, [scoring_row])
    write_csv(csv_trials, TRIAL_COLUMNS, trial_rows)
    write_csv(csv_pubs, PUB_COLUMNS, pub_rows)

    log.ok(f"CSV (scoring)      -> {csv_scoring}")
    log.ok(f"CSV (trials)       -> {csv_trials}")
    log.ok(f"CSV (publications) -> {csv_pubs}")

    if args_write_bq:
        log.sub("Writing to BigQuery")
        write_bq_to_bigquery(scoring_row, trial_rows, pub_rows)

    # Route the final report through the buffered logger too, so it's part
    # of the single atomic flush for this molecule rather than an
    # unbuffered print() that could interleave with other molecules.
    Log.raw(report)

    elapsed = time.time() - t0
    log.header(f"COMPLETE -- {elapsed:.1f}s elapsed")
    log.info(f"Files in: {OUTPUT_DIR.resolve()}")

    # Extract clean LLM consistency narrative directly from the stored detail
    llm_detail = getattr(score_consistency, '_last_llm_detail', None)
    if llm_detail:
        pct = llm_detail.get("pct", 0)
        summary = llm_detail.get("summary", "")
        llm_consistency_text = f"~{pct:.1f}% of {len(trials)} studies consistent positive.\n{summary}".strip()
    else:
        # Rule-based was used — surface the C reasoning as the report text
        llm_consistency_text = reasoning.get("C", "").replace("[Rule-based]", "").strip()

    return {
        "molecule":               molecule["generic_name"],
        "brand_names":            ", ".join(molecule.get("brand_names", [])),
        "sponsor":                molecule.get("sponsor", ""),
        "P":                      P,
        "D":                      D,
        "G":                      G,
        "E":                      E,
        "C":                      C,
        "Final Score":            final,
        "Evidence Level":         assessment.get("evidence_level", ""),
        "Total Trials":           len(trials),
        "RCT Count":              sum(1 for t in trials if "randomized" in str(t.get("study_design","")).lower()),
        "Publications":           len(publications),
        "FDA Approved":           "Yes" if assessment.get("regulatory_approval") else "No",
        "Evidence Gaps":          "; ".join(gaps),
        "Consistency":            consolidation.get("findings_consistency", ""),
        "LLM Consistency Report": llm_consistency_text,
    }


def write_pdgec_excel(rows: list, path: Path):
    """Write a summary Excel file with PDGEC scores for all molecules."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDGEC Scores"

    # Color palette
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
    SUB_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
    ALT_FILL  = PatternFill("solid", fgColor="D9E1F2")   # light blue
    GREEN_FILL= PatternFill("solid", fgColor="E2EFDA")
    RED_FILL  = PatternFill("solid", fgColor="FFE0E0")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SUB_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BOLD      = Font(bold=True, size=10)
    NORMAL    = Font(size=10)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="AAAAAA")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(r, c, value, font=None, fill=None, align=None, border=True):
        _c = ws.cell(row=r, column=c, value=value)
        if font:   _c.font   = font
        if fill:   _c.fill   = fill
        if align:  _c.alignment = align
        if border: _c.border = BORDER
        return _c

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    cell(1, 1, "PDGEC Evidence Score Summary", HDR_FONT, HDR_FILL, CENTER)
    ws.row_dimensions[1].height = 28

    # Section header row
    headers = [
        "Molecule", "Brand Names", "Sponsor",
        "P (Phase)", "D (Design)", "G (Geo)", "E (Evidence)", "C (Consistency)", "Final Score",
        "Evidence Level", "Total Trials", "RCT Count", "Publications", "FDA Approved"
    ]
    for col, h in enumerate(headers, 1):
        cell(2, col, h, SUB_FONT, SUB_FILL, CENTER)
    ws.row_dimensions[2].height = 36

    # Score columns (D-L = cols 4-9)
    SCORE_COLS = {4: "P", 5: "D", 6: "G", 7: "E", 8: "C", 9: "Final Score"}

    def score_fill(val, is_c=False):
        try:
            v = float(val)
        except (TypeError, ValueError):
            return None
        if is_c:
            return GREEN_FILL if v >= 0 else RED_FILL
        return GREEN_FILL if v >= 4 else (RED_FILL if v <= 2 else None)

    for row_idx, row in enumerate(rows, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        values = [
            row.get("molecule", ""),
            row.get("brand_names", ""),
            row.get("sponsor", ""),
            row.get("P", ""),
            row.get("D", ""),
            row.get("G", ""),
            row.get("E", ""),
            row.get("C", ""),
            row.get("Final Score", ""),
            row.get("Evidence Level", ""),
            row.get("Total Trials", ""),
            row.get("RCT Count", ""),
            row.get("Publications", ""),
            row.get("FDA Approved", ""),
        ]
        for col, val in enumerate(values, 1):
            is_c_col = (col == 8)
            is_score = col in SCORE_COLS
            sf = score_fill(val, is_c=is_c_col) if is_score else fill
            f = BOLD if col == 1 else NORMAL
            al = CENTER if col >= 4 else LEFT
            cell(row_idx, col, val, f, sf or fill, al)
        ws.row_dimensions[row_idx].height = 20

    # Column widths
    col_widths = [22, 28, 24, 9, 9, 9, 9, 9, 11, 16, 12, 10, 13, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Second sheet: Consistency & Evidence Gaps detail
    ws2 = wb.create_sheet("Detail")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _c = ws2.cell(row=1, column=1, value="Evidence Detail")
    _c.font = HDR_FONT; _c.fill = HDR_FILL; _c.alignment = CENTER
    ws2.row_dimensions[1].height = 24

    hdr2 = ["Molecule", "Consistency (rule-based)", "LLM Consistency Report", "Evidence Gaps"]
    for col, h in enumerate(hdr2, 1):
        _c = ws2.cell(row=2, column=col, value=h)
        _c.font = SUB_FONT; _c.fill = SUB_FILL; _c.alignment = CENTER; _c.border = BORDER
    ws2.row_dimensions[2].height = 30

    for row_idx, row in enumerate(rows, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        vals = [
            row.get("molecule", ""),
            row.get("Consistency", ""),
            row.get("LLM Consistency Report", ""),
            row.get("Evidence Gaps", ""),
        ]
        for col, val in enumerate(vals, 1):
            _c = ws2.cell(row=row_idx, column=col, value=val)
            _c.font = NORMAL
            _c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if fill: _c.fill = fill
            _c.border = BORDER
        # Taller rows to accommodate the LLM narrative
        ws2.row_dimensions[row_idx].height = 120

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 80
    ws2.column_dimensions["D"].width = 45

    wb.save(path)


# Default molecule list — runs when no --molecule argument is supplied
# Fallback molecule list — only used if BigQuery is unreachable
DEFAULT_MOLECULES = [
    "semaglutide", "liraglutide", "tirzepatide", "dulaglutide",
    "exenatide", "albiglutide", "insulin glargine",
    "empagliflozin", "dapagliflozin", "canagliflozin",
]


def main():
    parser = argparse.ArgumentParser(description="Clinical Evidence Analyzer v3")
    parser.add_argument(
        "--molecule", "-m", required=False, default=None, nargs="+",
        help=(
            "One or more molecule names, space-separated or comma-separated. "
            "Examples: --molecule semaglutide liraglutide tirzepatide "
            "OR --molecule semaglutide,liraglutide. "
            "Omit to auto-discover all molecules from the BigQuery table."
        )
    )
    parser.add_argument("--max-trials", type=int, default=20)
    parser.add_argument("--write-bq",   action="store_true",
                        help="Write results back to BigQuery tables")
    parser.add_argument("--max-molecule-workers", type=int, default=4,
                        help="Max number of molecules to process concurrently (default: 4)")
    args = parser.parse_args()

    # Gate: Gemini must be initialised before any processing starts
    if _gemini_model is None:
        log.err("Cannot run: Gemini API is not initialised.")
        log.err("Ensure GEMINI_API_KEY is set in your .env file and google-generativeai is installed.")
        sys.exit(1)

    if args.molecule:
        # Flatten space-separated tokens and handle any comma-separated values within them
        molecules = []
        for token in args.molecule:
            for m in token.split(","):
                m = m.strip()
                if m:
                    molecules.append(m)
        log.info(f"Running {len(molecules)} molecule(s) from --molecule flag.")
    else:
        log.info("No --molecule supplied. Fetching molecule list from BigQuery table...")
        molecules = fetch_bq_molecules()
        if not molecules:
            log.warn("BigQuery returned no molecules. Falling back to DEFAULT_MOLECULES list.")
            molecules = DEFAULT_MOLECULES
        log.info(f"Will process {len(molecules)} molecule(s): {', '.join(molecules[:10])}{'...' if len(molecules) > 10 else ''}")

    all_scores = []

    max_workers = min(args.max_molecule_workers, len(molecules)) or 1
    log.info(f"Processing molecules in parallel (max_workers={max_workers}, "
              f"gemini_max_concurrent={GEMINI_MAX_CONCURRENT})...")

    def _process_molecule(mol):
        """
        Runs one molecule end-to-end in a worker thread. Sets up a fresh
        per-molecule log buffer BEFORE any work starts, so every log.*
        call made by run_molecule() (and everything it calls, including
        its own nested thread pools via submit_with_ctx) writes into this
        buffer instead of printing directly. Once the molecule is fully
        done (success or failure), the buffer is flushed in one atomic
        print() call, so this molecule's full output appears as a single
        unbroken block even though other molecules are printing at the
        same time from other threads.
        """
        buf = io.StringIO()
        token = _log_buffer_var.set(buf)
        try:
            result = run_molecule(mol, args.max_trials, args.write_bq)
        except Exception as e:
            buf.write(f"    [ERR]  Failed to process {mol}: {e}\n")
            result = {"molecule": mol, "error": str(e)}
        finally:
            _log_buffer_var.reset(token)
            # Single atomic write of this molecule's entire log output.
            sys.stdout.write(buf.getvalue())
            sys.stdout.flush()
        return result

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_process_molecule, mol): mol for mol in molecules}
        for future in as_completed(futures):
            mol = futures[future]
            all_scores.append(future.result())

    ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = OUTPUT_DIR / f"pdgec_scores_{ts_file}.xlsx"
    write_pdgec_excel(all_scores, excel_path)
    log.ok(f"Excel summary -> {excel_path}")

if __name__ == "__main__":
    main()
